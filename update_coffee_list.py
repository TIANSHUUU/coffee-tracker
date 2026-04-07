#!/usr/bin/env python3
import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from html import unescape
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill


TIMEOUT_SECONDS = 25
DEFAULT_CONFIG = "config/roasters.json"
DEFAULT_OUTPUT_DIR = "output"
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/137.0.0.0 Safari/537.36"
)

COUNTRY_KEYWORDS = {
    "ethiopia",
    "kenya",
    "colombia",
    "brazil",
    "guatemala",
    "rwanda",
    "burundi",
    "indonesia",
    "panama",
    "el salvador",
    "honduras",
    "peru",
    "mexico",
    "uganda",
    "ecuador",
    "nicaragua",
    "costa rica",
    "bolivia",
    "yemen",
    "china",
    "tanzania",
    "papua new guinea",
    "png",
    "dr congo",
    "congo",
}

PROCESS_KEYWORDS = [
    "washed",
    "natural",
    "honey",
    "anaerobic",
    "wet hulled",
    "carbonic maceration",
    "double fermentation",
    "co-ferment",
    "experimental",
]

EXCLUDED_PRODUCT_KEYWORDS = [
    "drip bag",
    "drip bags",
    "bundle",
    "bundles",
    "subscription",
    "subsrcription",
    "equipment",
    "merchandise",
    "merch",
    "cup",
    "cascara tea",
    "decaf",
    "tote",
    "machine",
    "grinder",
    "kettle",
    "dripper",
    "filter paper",
    "filter papers",
    "paper filter",
    "paper filters",
    "magazine",
    "book",
    "gift card",
    "giftcard",
    "tea",
    "brewing scales",
    "scale",
    "scales",
    "candle",
    "jug",
    "jugs",
    "tee",
    "vest",
    "jumper",
    "matcha",
    "matcha latte",
    "cap",
    "tamper",
    "hario v60 drip assist set",
    "instant coffee",
    # Clothing & accessories
    "hoodie",
    "beanie",
    "apron",
    "socks",
    # Equipment & tools
    "bottle",
    "portafilter",
    "refractometer",
    "mug",
    "glasses",
    "pitcher",
    # Food & non-coffee products
    "panela sugar",
    "panel sugar",
    "sampler pack",
]

EXCLUDED_PRODUCT_REGEXES = [
    r"\bcoffee\s*concentrate\b",
    r"\bt\s*shirts?\b",
    r"\btee\s*shirts?\b",
    r"\bgift\s*cards?\b",
    r"\bncd\b",       # NCD / NCD Pulse / Pink NCD refractometer
    r"\bstem\b",      # STEM milk pitcher
]

ROAST_PROFILE_COLORS = {
    "filter": "pink",
    "espresso": "green",
    "omni": "blue",
}


@dataclass
class CoffeeItem:
    roaster: str
    source_url: str
    bean_name: str
    roast_profile: str
    origin: str
    price_aud: str
    process: str
    varietal: str
    flavour_profile: str
    product_url: str
    status: str
    error: str


def clean_text(text: str) -> str:
    text = unescape(text or "")
    text = re.sub(r"<[^>]+>", " ", text)  # strip HTML tags
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_host(host: str) -> str:
    return (host or "").lower().strip().removeprefix("www.")


def same_site(url_a: str, url_b: str) -> bool:
    return normalize_host(urlparse(url_a).netloc) == normalize_host(urlparse(url_b).netloc)


def alternate_www_url(url: str) -> Optional[str]:
    parsed = urlparse(url)
    host = parsed.netloc
    if not host:
        return None
    if host.startswith("www."):
        alt_host = host[4:]
    else:
        alt_host = f"www.{host}"
    rebuilt = parsed._replace(netloc=alt_host)
    return rebuilt.geturl()


def unescape_url_blob(text: str) -> str:
    # Handle JSON-style escaped URLs from script blobs.
    return (
        text.replace("\\/", "/")
        .replace("\\u002F", "/")
        .replace("\\u002f", "/")
        .replace("\\u003A", ":")
        .replace("\\u003a", ":")
    )


def normalize_for_matching(text: str) -> str:
    # Keep only alphanumeric tokens to make matching resilient to punctuation/hyphen variants.
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", clean_text(text).lower())).strip()


def extract_price(text: str) -> str:
    m = re.search(r"\$\s*([0-9]+(?:\.[0-9]{1,2})?)", text)
    if not m:
        return ""
    return m.group(1)


def parse_roast_profile(text: str, title: str = "") -> str:
    # Title is the most reliable signal — if it explicitly says filter/espresso, trust it.
    if title:
        t_title = title.lower()
        title_filter   = bool(re.search(r"\bfilter\b", t_title))
        title_espresso = bool(re.search(r"\bespresso\b", t_title))
        title_omni     = bool(re.search(r"\bomni\b", t_title))
        if title_omni or (title_filter and title_espresso):
            return "omni"
        if title_filter:
            return "filter"
        if title_espresso:
            return "espresso"

    t = text.lower()
    has_omni = bool(re.search(r"\bomni(?:\s*roast)?\b", t))
    has_filter = bool(re.search(r"\b(filter|pour over|pourover|v60|batch brew)\b", t))
    has_espresso = bool(re.search(r"\bespresso\b", t))

    if has_omni or (has_filter and has_espresso):
        return "omni"
    if has_filter:
        return "filter"
    if has_espresso:
        return "espresso"
    return ""


def parse_origin(text: str) -> str:
    t = clean_text(text)
    lower = t.lower()

    # "origin/country:" labels are reliable; "from/grown in" is too broad so we
    # only trust it when the captured text contains a known country keyword.
    explicit = re.search(r"(?:origin|country)\s*[:\-]\s*([A-Za-z ,/&]+)", t, flags=re.IGNORECASE)
    if explicit:
        candidate = clean_text(explicit.group(1))
        candidate = re.split(r"[.|;]| process| variety| roast", candidate, flags=re.IGNORECASE)[0]
        return candidate[:80].strip(" ,")

    grown_in = re.search(r"(?:from|grown in)\s+([A-Za-z ,/&]+)", t, flags=re.IGNORECASE)
    if grown_in:
        candidate = clean_text(grown_in.group(1))
        candidate = re.split(r"[.|;]| process| variety| roast", candidate, flags=re.IGNORECASE)[0]
        candidate = candidate[:80].strip(" ,")
        # Only trust if it actually contains a country name
        if any(country in candidate.lower() for country in COUNTRY_KEYWORDS):
            return candidate

    found = []
    for country in COUNTRY_KEYWORDS:
        if country in lower:
            found.append(country.upper() if country == "png" else country.title())
    if found:
        return ", ".join(sorted(set(found)))

    return ""


def parse_process(text: str, title: str = "", product_url: str = "") -> str:
    t = clean_text(text)

    def detect_process_terms(blob: str) -> List[str]:
        term_patterns = [
            (r"\bwashed\b", "Washed"),
            (r"\bnatural\b", "Natural"),
            (r"\bhoney\b", "Honey"),
            (r"\banaerobic\b", "Anaerobic"),
            (r"\bwet\s+hulled\b", "Wet Hulled"),
            (r"\bcarbonic\s+maceration\b", "Carbonic Maceration"),
            (r"\bdouble\s+fermentation\b", "Double Fermentation"),
            (r"\bco[\s-]?ferment\b", "Co-Ferment"),
            (r"\bexperimental\b", "Experimental"),
        ]
        found: List[str] = []
        for pattern, label in term_patterns:
            if re.search(pattern, blob, flags=re.IGNORECASE):
                found.append(label)
        return list(dict.fromkeys(found))

    explicit = re.search(r"process(?:ing)?\s*[:\-]\s*([A-Za-z ,/&-]+)", t, flags=re.IGNORECASE)
    if explicit:
        candidate = clean_text(explicit.group(1))
        candidate = re.split(r"[.|;]| origin| variety| roast", candidate, flags=re.IGNORECASE)[0]
        return candidate[:80].strip(" ,")

    for priority_blob in [title, urlparse(product_url).path.replace("-", " ")]:
        matches = detect_process_terms(clean_text(priority_blob))
        if matches:
            return ", ".join(matches)

    matches = detect_process_terms(t)
    if matches:
        return ", ".join(matches)

    return ""


def parse_varietal(text: str) -> str:
    t = clean_text(text)
    patterns = [
        r"(?:variet(?:al|y|ies)|cultivar)\s*[:\-]\s*([A-Za-z0-9 ,/&()+-]+)",
    ]
    for p in patterns:
        m = re.search(p, t, flags=re.IGNORECASE)
        if m:
            candidate = clean_text(m.group(1))
            candidate = re.split(r"[.;]| origin| process| roast| altitude| notes?", candidate, flags=re.IGNORECASE)[0]
            return candidate[:120].strip(" ,")
    return ""


def parse_flavour_profile(text: str) -> str:
    t = clean_text(text)

    stop_labels = (
        r"(?:"
        r"origin|country|process(?:ing)?|variet(?:y|al|ies)|cultivar|roast|region|producer|farm|"
        r"altitude|method|weight|size|brew|grind|component|price|quantity|"
        r"add\s+to\s+cart|subscription"
        r")"
    )

    flavour_hints = [
        "citrus",
        "orange",
        "lemon",
        "lime",
        "grapefruit",
        "tangerine",
        "bergamot",
        "berry",
        "strawberry",
        "raspberry",
        "blueberry",
        "cherry",
        "stone fruit",
        "peach",
        "apricot",
        "plum",
        "nectarine",
        "tropical",
        "pineapple",
        "mango",
        "passionfruit",
        "floral",
        "jasmine",
        "tea",
        "chocolate",
        "cocoa",
        "caramel",
        "toffee",
        "honey",
        "vanilla",
        "marzipan",
        "sugarcane",
        "almond",
        "hazelnut",
        "blackcurrant",
        "herbal",
        "spice",
    ]

    def looks_like_flavour(candidate: str) -> bool:
        lower = candidate.lower()
        # Reject common false-positive fragments.
        if any(
            phrase in lower
            for phrase in [
                "profile is",
                "easy to",
                "top eight coffees",
                "to your sweet creations",
                "competition",
            ]
        ):
            return False
        if any(hint in lower for hint in flavour_hints):
            return True
        # Keep short comma-separated note lists even when hints are sparse.
        if "," in candidate and len(candidate.split()) <= 14:
            return True
        return False

    def clean_candidate(raw: str) -> str:
        candidate = clean_text(raw)
        candidate = re.sub(rf"\s+(?={stop_labels}\s*[:\-])", " | ", candidate, flags=re.IGNORECASE)
        candidate = candidate.split(" | ", 1)[0]
        candidate = re.split(
            r"\b(?:v60\s+filter\s+recipe|espresso\s+recipe|recipe|dose|yield|time)\b",
            candidate,
            flags=re.IGNORECASE,
        )[0]
        candidate = re.split(r"\broast\b", candidate, flags=re.IGNORECASE)[0]
        candidate = re.split(r"[.;]| \| ", candidate, flags=re.IGNORECASE)[0]
        candidate = candidate.strip(" ,:-")
        if len(candidate) < 6:
            return ""
        if len(candidate.split()) > 20:
            return ""
        if not looks_like_flavour(candidate):
            return ""
        return candidate[:160]

    patterns = [
        # Common style: CUP: CITRUS, SUGARCANE, SWEET HERBALS
        r"\bcup(?:\s*profile)?\b\s*[:\-]\s*([A-Za-z0-9' ,/&()+-]{6,180})",
        # Common style: TASTES LIKE Tropical fruit, marzipan and dark chocolate
        r"\btastes?\s+like\b\s*[:\-]?\s*([A-Za-z0-9' ,/&()+-]{6,180})",
        # Generic styles: flavour/flavor/tasting notes/notes
        r"(?:flavour|flavor|tasting\s*notes?|notes?)\s*[:\-]\s*([A-Za-z0-9' ,/&()+-]{6,180})",
    ]
    for p in patterns:
        m = re.search(p, t, flags=re.IGNORECASE)
        if not m:
            continue
        candidate = clean_candidate(m.group(1))
        if candidate:
            return candidate

    # Fallback for label without colon: "CUP citrus, sugarcane, sweet herbals"
    fallback = re.search(r"\bcup\b\s+([A-Za-z0-9' ,/&()+-]{8,140})", t, flags=re.IGNORECASE)
    if fallback:
        candidate = clean_candidate(fallback.group(1))
        if candidate and ("," in candidate or " and " in candidate.lower()):
            return candidate

    return ""


def should_exclude_product(text: str) -> bool:
    normalized = normalize_for_matching(text)
    for keyword in EXCLUDED_PRODUCT_KEYWORDS:
        pattern = r"\b" + r"\s+".join(re.escape(p) for p in keyword.split()) + r"\b"
        if re.search(pattern, normalized):
            return True
    return any(re.search(pattern, normalized) for pattern in EXCLUDED_PRODUCT_REGEXES)


def should_exclude_product_for_roaster(roaster: str, text: str) -> bool:
    if should_exclude_product(text):
        return True
    if roaster.strip().lower() == "puchero coffee":
        lower = clean_text(text).lower()
        if re.search(r"\bchocolate\b", lower):
            return True
        if re.search(r"\bt[eé]\b", lower):
            return True
    return False


def get_roast_profile_color(roast_profile: str) -> str:
    return ROAST_PROFILE_COLORS.get(roast_profile, "")


def session_with_headers() -> requests.Session:
    s = requests.Session()
    s.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-AU,en;q=0.9,en-US;q=0.8",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
            "Referer": "https://www.google.com/",
        }
    )
    return s


def normalize_collection_json_url(
    listing_url: str, shopify_collection_handle: Optional[str] = None
) -> Optional[str]:
    parsed = urlparse(listing_url)
    if shopify_collection_handle:
        handle = shopify_collection_handle.strip("/")
        if handle:
            return f"{parsed.scheme}://{parsed.netloc}/collections/{handle}/products.json?limit=250&country=AU"

    parts = [p for p in parsed.path.split("/") if p]
    if len(parts) >= 2 and parts[0] == "collections":
        handle = parts[1]
        return f"{parsed.scheme}://{parsed.netloc}/collections/{handle}/products.json?limit=250&country=AU"
    return None


# Regex to detect JS-template placeholder labels captured by accident
# e.g. "IMPORTER/PARTNER", "PROCESSING", "REGION", "GEO TAG"
_TEMPLATE_LABEL_RE = re.compile(r"^[A-Z][A-Z0-9 /]+$")

def _is_template_label(s: str) -> bool:
    """Return True if s looks like an all-caps template label, not a real value."""
    return bool(_TEMPLATE_LABEL_RE.match(s.strip()))


def extract_shopify_tag_metadata(tags: List[str]) -> Dict[str, List[str]]:
    """Parse structured Shopify tags like 'PROCESS: Washed' into a metadata dict."""
    meta: Dict[str, List[str]] = {
        "process": [], "country": [], "roast": [], "varietal": [], "producer": [], "region": []
    }
    for tag in (tags or []):
        # Format 1: "ROAST: Filter" / "PROCESS: Washed" etc.
        m = re.match(r"^(PROCESS|COUNTRY|ROAST|VARIETAL|PRODUCER|REGION)\s*:\s*(.+)$",
                     tag.strip(), re.IGNORECASE)
        if m:
            key = m.group(1).lower()
            val = m.group(2).strip()
            if key in meta and val:
                meta[key].append(val)
            continue
        # Format 2: "Coffee Type.Filter" / "Coffee Type.Espresso" — definitive ONA tag
        m2 = re.match(r"^Coffee\s+Type[.\s]+(.+)$", tag.strip(), re.IGNORECASE)
        if m2:
            ct = m2.group(1).strip().lower()
            if re.search(r"\bfilter\b", ct):
                meta["_coffee_type"] = "filter"
            elif re.search(r"\bespresso\b", ct):
                meta["_coffee_type"] = "espresso"
            continue
        # Format 3: "Brew Method.Espresso" / "Brew Method.Pour over" (ONA-style)
        m3 = re.match(r"^Brew\s+Method[.\s]+(.+)$", tag.strip(), re.IGNORECASE)
        if m3:
            brew = m3.group(1).strip().lower()
            if re.search(r"\b(filter|pour.?over|v60|aeropress|batch brew|drip)\b", brew):
                meta["roast"].append("filter")
            elif re.search(r"\bespresso\b", brew):
                meta["roast"].append("espresso")
    # Coffee Type overrides Brew Method — it's the explicit roast classification
    if "_coffee_type" in meta:
        meta["roast"] = [meta.pop("_coffee_type")]
    return meta



def scrape_shopify_collection_json(
    session: requests.Session,
    roaster: str,
    listing_url: str,
    shopify_collection_handle: Optional[str] = None,
    price_variant_filter: Optional[str] = None,
) -> Tuple[List[CoffeeItem], Optional[str]]:
    json_url = normalize_collection_json_url(listing_url, shopify_collection_handle)
    if not json_url:
        return [], "not_a_collection_url"

    try:
        resp = session.get(json_url, timeout=TIMEOUT_SECONDS)
        if resp.status_code != 200:
            return [], f"shopify_json_http_{resp.status_code}"
        payload = resp.json()
    except Exception as exc:
        return [], f"shopify_json_error: {exc}"

    products = payload.get("products") or []
    if not products:
        return [], "shopify_json_no_products"

    parsed = urlparse(listing_url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    items: List[CoffeeItem] = []

    for product in products:
        title = clean_text(str(product.get("title", "")))
        if not title:
            continue

        raw_tags: List[str] = product.get("tags") or []
        tag_meta = extract_shopify_tag_metadata(raw_tags)
        tags_str = " ".join(raw_tags)
        body_html = product.get("body_html") or ""
        text_blob = clean_text(BeautifulSoup(body_html, "html.parser").get_text(" "))
        combined = " ".join([title, tags_str, text_blob])
        if should_exclude_product_for_roaster(roaster, title):
            continue

        price = ""
        variants = product.get("variants") or []
        if variants:
            try:
                filtered = (
                    [v for v in variants if price_variant_filter and price_variant_filter.lower() in (v.get("title") or "").lower()]
                    or variants
                )
                price_values = [float(v.get("price")) for v in filtered if v.get("price") is not None]
                if price_values:
                    price = f"{min(price_values):.2f}"
            except Exception:
                price = str(variants[0].get("price") or "")

        handle = product.get("handle") or ""
        product_url = f"{base}/products/{handle}" if handle else listing_url

        # Use individual product.json for two purposes:
        # 1. Availability check (skip sold-out products)
        # 2. Price override — country=AU forces Shopify Markets to return AUD price.
        if handle:
            try:
                pj_resp = session.get(f"{base}/products/{handle}.json?country=AU", timeout=TIMEOUT_SECONDS)
                if pj_resp.status_code == 200:
                    pj = pj_resp.json().get("product", {})
                    # Availability: skip if no published_at (unpublished = sold out / unavailable)
                    pj_variants = pj.get("variants") or []
                    all_unavailable = pj_variants and all(
                        v.get("inventory_management") == "shopify" and
                        v.get("inventory_policy") == "deny" and
                        int(v.get("inventory_quantity") or 0) <= 0
                        for v in pj_variants
                    )
                    if all_unavailable:
                        continue
                    # Price: use AUD-verified price from product.json if currency matches
                    if pj_variants:
                        currency = pj_variants[0].get("price_currency", "")
                        if currency == "AUD":
                            pj_filtered = (
                                [v for v in pj_variants if price_variant_filter and price_variant_filter.lower() in (v.get("title") or "").lower()]
                                or pj_variants
                            )
                            pj_prices = [float(v["price"]) for v in pj_filtered if v.get("price")]
                            if pj_prices:
                                price = f"{min(pj_prices):.2f}"
            except Exception:
                pass  # If check fails, keep collection JSON price and include the product

        # Prefer structured tag values; fall back to body-text parsing.
        # Guard against JS-template placeholders captured as values (e.g. "PROCESSING", "REGION").
        def _tag_or_parse(tag_vals: List[str], fallback: str) -> str:
            if tag_vals:
                return ", ".join(tag_vals)
            return "" if _is_template_label(fallback) else fallback

        process = _tag_or_parse(
            tag_meta["process"],
            parse_process(combined, title=title, product_url=product_url),
        )
        origin = _tag_or_parse(
            tag_meta["country"],
            parse_origin(combined),
        )
        varietal = _tag_or_parse(
            tag_meta["varietal"],
            parse_varietal(combined),
        )
        # If variants offer multiple brew methods (espresso + filter), the product is omni-roast.
        variant_titles = " ".join(v.get("title") or "" for v in variants).lower()
        has_espresso_variant = any(k in variant_titles for k in ("espresso",))
        has_filter_variant = any(k in variant_titles for k in ("pour over", "filter", "batch brew", "aeropress", "plunger"))
        if has_espresso_variant and has_filter_variant:
            roast_profile_label = "omni"
        elif tag_meta["roast"]:
            roast_profile_label = parse_roast_profile(" ".join(tag_meta["roast"]), title=title)
        else:
            roast_profile_label = parse_roast_profile(combined, title=title)

        items.append(
            CoffeeItem(
                roaster=roaster,
                source_url=product_url,
                bean_name=title,
                roast_profile=roast_profile_label,
                origin=origin,
                price_aud=price,
                process=process,
                varietal=varietal,
                flavour_profile=parse_flavour_profile(combined),
                product_url=product_url,
                status="ok",
                error="",
            )
        )

    if not items:
        return [], "shopify_json_no_parsed_items"
    return items, None


def scrape_shopify_all_products_json(
    session: requests.Session, roaster: str, listing_url: str
) -> Tuple[List[CoffeeItem], Optional[str]]:
    parsed = urlparse(listing_url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    candidate_urls = [
        f"{base}/products.json?limit=250&country=AU",
        f"{base}/collections/all/products.json?limit=250&country=AU",
        f"{base}/collections/shop/products.json?limit=250&country=AU",
        f"{base}/collections/coffee/products.json?limit=250&country=AU",
    ]
    errors: List[str] = []
    best_items: List[CoffeeItem] = []

    for json_url in candidate_urls:
        try:
            resp = session.get(json_url, timeout=TIMEOUT_SECONDS)
            if resp.status_code != 200:
                errors.append(f"{json_url}:http_{resp.status_code}")
                continue
            payload = resp.json()
        except Exception as exc:
            errors.append(f"{json_url}:error:{exc}")
            continue

        products = payload.get("products") or []
        if not products:
            errors.append(f"{json_url}:no_products")
            continue

        items: List[CoffeeItem] = []
        for product in products:
            title = clean_text(str(product.get("title", "")))
            if not title:
                continue

            body_html = product.get("body_html") or ""
            tags = " ".join(product.get("tags") or [])
            text_blob = clean_text(BeautifulSoup(body_html, "html.parser").get_text(" "))
            exclude_text = " ".join([title, tags])
            combined = " ".join([title, tags, text_blob])
            if should_exclude_product_for_roaster(roaster, exclude_text):
                continue

            price = ""
            variants = product.get("variants") or []
            if variants:
                try:
                    price_values = [float(v.get("price")) for v in variants if v.get("price") is not None]
                    if price_values:
                        price = f"{min(price_values):.2f}"
                except Exception:
                    price = str(variants[0].get("price") or "")

            handle = product.get("handle") or ""
            product_url = f"{base}/products/{handle}" if handle else listing_url
            roast_profile = parse_roast_profile(combined)
            items.append(
                CoffeeItem(
                    roaster=roaster,
                    source_url=product_url,
                    bean_name=title,
                    roast_profile=roast_profile,
                    origin=parse_origin(combined),
                    price_aud=price,
                    process=parse_process(combined, title=title, product_url=product_url),
                    varietal=parse_varietal(combined),
                    flavour_profile=parse_flavour_profile(combined),
                    product_url=product_url,
                    status="ok",
                    error="",
                )
            )

        if len(items) > len(best_items):
            best_items = items

    if best_items:
        return best_items, None
    if errors:
        return [], "shopify_all_json_failed: " + "; ".join(errors[:4])
    return [], "shopify_all_json_no_results"


def discover_collection_handles_from_listing(
    session: requests.Session, listing_url: str
) -> Tuple[List[str], Optional[str]]:
    try:
        resp = session.get(listing_url, timeout=TIMEOUT_SECONDS)
        resp.raise_for_status()
    except Exception as exc:
        return [], f"handle_discovery_error: {exc}"

    handles = re.findall(r"/collections/([a-z0-9][a-z0-9-]*)", resp.text, flags=re.IGNORECASE)
    unique: List[str] = []
    seen = set()
    for h in handles:
        handle = h.lower().strip("-")
        if not handle or handle in {"all", "frontpage"} or handle in seen:
            continue
        seen.add(handle)
        unique.append(handle)

    if not unique:
        return [], "handle_discovery_no_candidates"

    def score(handle: str) -> int:
        s = 0
        if "coffee" in handle:
            s += 5
        if "single" in handle or "origin" in handle:
            s += 3
        if "bean" in handle:
            s += 2
        if "shop" in handle:
            s += 1
        return s

    ranked = sorted(unique, key=lambda h: (score(h), -len(h)), reverse=True)
    return ranked[:6], None


def normalize_product_url(base_url: str, candidate_url: str) -> Optional[str]:
    full = urljoin(base_url, candidate_url.strip())
    parsed = urlparse(full)
    if "/products/" not in parsed.path and "/product/" not in parsed.path:
        return None
    return f"{parsed.scheme}://{parsed.netloc}{parsed.path}"


def extract_product_links_from_html(base_url: str, html: str) -> List[str]:
    soup = BeautifulSoup(html, "html.parser")
    links: List[str] = []
    seen = set()

    def maybe_add(candidate: str) -> None:
        normalized = normalize_product_url(base_url, candidate)
        if normalized and normalized not in seen:
            seen.add(normalized)
            links.append(normalized)

    for a in soup.find_all("a", href=True):
        maybe_add(a["href"])

    for tag in soup.find_all(True):
        for attr in ("data-product-url", "data-url", "data-href", "href"):
            value = tag.attrs.get(attr)
            if isinstance(value, str) and ("/products/" in value or "/product/" in value):
                maybe_add(value)

    # Many storefronts render product URLs inside JSON/script blobs instead of anchor tags.
    def scan_blob(blob: str) -> None:
        for m in re.findall(r'["\'](/products?/[^"\']+)["\']', blob):
            maybe_add(m)
        for m in re.findall(r'https?://[^"\'>\s]+/products?/[^"\'>\s]+', blob):
            maybe_add(m)

    scan_blob(html)
    # Handle JSON-escaped URLs like https:\/\/example.com\/products\/slug.
    scan_blob(unescape_url_blob(html))

    return links


def extract_shop_slug_links(base_url: str, html: str) -> List[str]:
    parsed_base = urlparse(base_url)
    base_prefix = f"{parsed_base.scheme}://{parsed_base.netloc}"
    links: List[str] = []
    seen = set()

    def maybe_add(candidate: str) -> None:
        full = urljoin(base_url, candidate.strip())
        parsed = urlparse(full)
        if not same_site(base_url, full):
            return
        path = parsed.path.rstrip("/")
        # Handle Woo-style product permalinks:
        #   /shop/<slug>  or  /shop/<category>/<slug>
        m = re.match(r"^/shop/((?:[^/]+/)?[^/]+)$", path)
        if not m:
            return
        slug = m.group(1).lower()
        first_segment = slug.split("/")[0]
        if first_segment in {"shop", "page", "category", "tag", "feed", "cart", "checkout", "my-account"}:
            return
        normalized = f"{base_prefix}/shop/{m.group(1)}"
        if normalized not in seen:
            seen.add(normalized)
            links.append(normalized)

    def scan_blob(blob: str) -> None:
        for m in re.findall(r'https?://[^"\'>\s]+/shop/[^"\'>\s]+', blob):
            maybe_add(m)
        for m in re.findall(r'["\'](/shop/[^"\']+)["\']', blob):
            maybe_add(m)

    scan_blob(html)
    # Handle JSON-escaped URLs like https:\/\/www.smallbatch.com.au\/shop\/slug
    scan_blob(unescape_url_blob(html))

    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        maybe_add(a["href"])

    return links


def parse_product_page(
    session: requests.Session, roaster: str, source_url: str, product_url: str
) -> CoffeeItem:
    try:
        resp = session.get(product_url, timeout=TIMEOUT_SECONDS)
        resp.raise_for_status()
    except Exception as exc:
        return CoffeeItem(
            roaster=roaster,
            source_url=source_url,
            bean_name="",
            roast_profile="",
            origin="",
            price_aud="",
            process="",
            varietal="",
            flavour_profile="",
            product_url=product_url,
            status="error",
            error=f"product_page_error: {exc}",
        )

    content_type = (resp.headers.get("Content-Type") or "").lower()
    if "xml" in content_type or "rss" in content_type or "atom" in content_type:
        return CoffeeItem(
            roaster=roaster,
            source_url=product_url,
            bean_name="",
            roast_profile="",
            origin="",
            price_aud="",
            process="",
            varietal="",
            flavour_profile="",
            product_url=product_url,
            status="skip",
            error="non_product_feed_page",
        )

    soup = BeautifulSoup(resp.text, "html.parser")

    title = ""
    og_title = soup.find("meta", attrs={"property": "og:title"})
    if og_title and og_title.get("content"):
        title = clean_text(og_title["content"])
    if not title and soup.find("h1"):
        title = clean_text(soup.find("h1").get_text(" "))

    price = ""
    og_price = soup.find("meta", attrs={"property": "product:price:amount"})
    if og_price and og_price.get("content"):
        price = clean_text(og_price["content"])
    if not price:
        ld = soup.find("script", attrs={"type": "application/ld+json"})
        if ld and ld.string:
            text = ld.string
            m = re.search(r'"price"\s*:\s*"?([0-9]+(?:\.[0-9]+)?)"?', text)
            if m:
                price = m.group(1)
    if not price:
        # Try WooCommerce / structured price element before scanning the full page.
        # Full-page scan risks picking up cart totals ($0) that appear early in the DOM.
        for sel in (".price", ".woocommerce-Price-amount", "[class*='price']"):
            el = soup.select_one(sel)
            if el:
                price = extract_price(el.get_text(" "))
                if price and price != "0":
                    break
    if not price or price == "0":
        price = extract_price(soup.get_text(" "))

    text_blob = clean_text(soup.get_text(" "))
    exclude_text = " ".join([title, product_url])
    if should_exclude_product_for_roaster(roaster, exclude_text):
        return CoffeeItem(
            roaster=roaster,
            source_url=product_url,
            bean_name=title,
            roast_profile="",
            origin="",
            price_aud=price,
            process="",
            varietal="",
            flavour_profile="",
            product_url=product_url,
            status="skip",
            error="excluded_non_beans",
        )

    roast_profile = parse_roast_profile(text_blob)

    return CoffeeItem(
        roaster=roaster,
        source_url=product_url,
        bean_name=title,
        roast_profile=roast_profile,
        origin=parse_origin(text_blob),
        price_aud=price,
        process=parse_process(text_blob, title=title, product_url=product_url),
        varietal=parse_varietal(text_blob),
        flavour_profile=parse_flavour_profile(text_blob),
        product_url=product_url,
        status="ok" if title else "error",
        error="" if title else "missing_product_title",
    )


def scrape_via_html_listing(
    session: requests.Session, roaster: str, listing_url: str
) -> Tuple[List[CoffeeItem], Optional[str]]:
    try:
        resp = session.get(listing_url, timeout=TIMEOUT_SECONDS)
        resp.raise_for_status()
    except Exception as exc:
        return [], f"listing_page_error: {exc}"

    product_links = extract_product_links_from_html(listing_url, resp.text)
    product_links = [
        url
        for url in product_links
        if not should_exclude_product_for_roaster(
            roaster, urlparse(url).path.replace("/", " ").replace("-", " ")
        )
    ]
    if not product_links:
        return [], "listing_page_no_product_links"

    items = [parse_product_page(session, roaster, listing_url, url) for url in product_links]
    items = [i for i in items if i.status != "skip"]
    ok_items = [i for i in items if i.status == "ok"]
    if not ok_items:
        return items, "listing_page_no_valid_products"

    return items, None


def scrape_via_shop_slug_listing(
    session: requests.Session, roaster: str, listing_url: str
) -> Tuple[List[CoffeeItem], Optional[str]]:
    try:
        resp = session.get(listing_url, timeout=TIMEOUT_SECONDS)
        resp.raise_for_status()
    except Exception as exc:
        return [], f"shop_slug_listing_error: {exc}"

    product_links = extract_shop_slug_links(listing_url, resp.text)
    product_links = [url for url in product_links if "/feed" not in urlparse(url).path.lower()]
    if not product_links:
        return [], "shop_slug_listing_no_product_links"

    items = [parse_product_page(session, roaster, listing_url, url) for url in product_links]
    items = [i for i in items if i.status != "skip"]
    ok_items = [i for i in items if i.status == "ok"]
    if not ok_items:
        return items, "shop_slug_listing_no_valid_products"

    return items, None


def extract_locs_from_sitemap_xml(xml_text: str) -> List[str]:
    return [clean_text(loc) for loc in re.findall(r"<loc>(.*?)</loc>", xml_text, flags=re.IGNORECASE | re.DOTALL)]


def discover_sitemaps_from_robots(session: requests.Session, base: str) -> List[str]:
    robots_url = f"{base}/robots.txt"
    try:
        resp = session.get(robots_url, timeout=TIMEOUT_SECONDS)
        if resp.status_code != 200:
            return []
    except Exception:
        return []

    sitemaps: List[str] = []
    for line in resp.text.splitlines():
        m = re.match(r"\s*Sitemap:\s*(\S+)\s*$", line, flags=re.IGNORECASE)
        if not m:
            continue
        sitemaps.append(clean_text(m.group(1)))
    return sitemaps


def scrape_via_sitemap(
    session: requests.Session, roaster: str, listing_url: str
) -> Tuple[List[CoffeeItem], Optional[str]]:
    parsed = urlparse(listing_url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    sitemap_candidates = [
        f"{base}/sitemap.xml",
        f"{base}/sitemap_index.xml",
        f"{base}/wp-sitemap.xml",
        f"{base}/product-sitemap.xml",
        f"{base}/product-sitemap1.xml",
        f"{base}/product-sitemap2.xml",
    ]
    sitemap_candidates.extend(discover_sitemaps_from_robots(session, base))

    discovered_sitemaps: List[str] = []
    seen_sitemaps = set()
    product_urls: List[str] = []
    seen_products = set()
    errors: List[str] = []

    def add_product_url(url: str) -> None:
        normalized = url.strip()
        p = urlparse(normalized)
        if not same_site(listing_url, normalized):
            return
        path = p.path.lower()
        if "/feed" in path:
            return
        if "/products/" not in path and "/product/" not in path and "/shop/" not in path:
            return
        if normalized not in seen_products:
            seen_products.add(normalized)
            product_urls.append(normalized)

    def fetch_sitemap(url: str) -> None:
        try:
            resp = session.get(url, timeout=TIMEOUT_SECONDS)
            if resp.status_code != 200:
                errors.append(f"sitemap_http_{resp.status_code}:{url}")
                return
            locs = extract_locs_from_sitemap_xml(resp.text)
            if not locs:
                errors.append(f"sitemap_no_locs:{url}")
                return
            for loc in locs:
                loc_lower = loc.lower()
                if loc_lower.endswith(".xml") or "sitemap" in loc_lower:
                    if loc not in seen_sitemaps:
                        seen_sitemaps.add(loc)
                        discovered_sitemaps.append(loc)
                else:
                    add_product_url(loc)
        except Exception as exc:
            errors.append(f"sitemap_error:{url}:{exc}")

    for sm in sitemap_candidates:
        if sm not in seen_sitemaps:
            seen_sitemaps.add(sm)
            discovered_sitemaps.append(sm)

    # Crawl discovered sitemap files with a bounded breadth-first pass.
    idx = 0
    while idx < len(discovered_sitemaps) and idx < 20:
        fetch_sitemap(discovered_sitemaps[idx])
        idx += 1

    if not product_urls:
        return [], "sitemap_no_product_urls"

    items = [parse_product_page(session, roaster, listing_url, url) for url in product_urls[:300]]
    items = [i for i in items if i.status != "skip"]
    ok_items = [i for i in items if i.status == "ok"]
    if not ok_items:
        detail = "; ".join(errors[:5])
        return items, f"sitemap_no_valid_products{(': ' + detail) if detail else ''}"

    return items, None


def scrape_woocommerce_store_api(
    session: requests.Session, roaster: str, listing_url: str
) -> Tuple[List[CoffeeItem], Optional[str]]:
    parsed = urlparse(listing_url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    endpoint = f"{base}/wp-json/wc/store/v1/products"

    all_products: List[Dict[str, object]] = []
    last_err: Optional[str] = None

    for page in range(1, 6):
        try:
            resp = session.get(
                endpoint,
                params={"per_page": 100, "page": page},
                timeout=TIMEOUT_SECONDS,
            )
            if resp.status_code != 200:
                last_err = f"woocommerce_http_{resp.status_code}"
                break
            payload = resp.json()
            if not isinstance(payload, list):
                return [], "woocommerce_invalid_payload"
            if not payload:
                break
            all_products.extend(payload)
            if len(payload) < 100:
                break
        except Exception as exc:
            return [], f"woocommerce_error: {exc}"

    if not all_products:
        return [], last_err or "woocommerce_no_products"

    items: List[CoffeeItem] = []
    for product in all_products:
        title = clean_text(str(product.get("name", "")))
        permalink = clean_text(str(product.get("permalink", "")))
        if not title or not permalink:
            continue

        short_desc = clean_text(
            BeautifulSoup(str(product.get("short_description", "")), "html.parser").get_text(" ")
        )
        description = clean_text(
            BeautifulSoup(str(product.get("description", "")), "html.parser").get_text(" ")
        )
        categories = " ".join(
            clean_text(str(c.get("name", ""))) for c in (product.get("categories") or []) if isinstance(c, dict)
        )
        tags = " ".join(
            clean_text(str(t.get("name", ""))) for t in (product.get("tags") or []) if isinstance(t, dict)
        )

        exclude_text = " ".join([title, permalink, categories, tags])
        if should_exclude_product_for_roaster(roaster, exclude_text):
            continue

        price = ""
        prices = product.get("prices")
        if isinstance(prices, dict):
            raw_price = str(prices.get("price", "") or "").strip()
            try:
                minor = int(prices.get("currency_minor_unit", 2))
                if raw_price.isdigit():
                    price = f"{int(raw_price) / (10 ** minor):.2f}"
            except Exception:
                price = ""

        combined = " ".join([title, short_desc, description, categories, tags])
        roast_profile = parse_roast_profile(combined)
        items.append(
            CoffeeItem(
                roaster=roaster,
                source_url=permalink,
                bean_name=title,
                roast_profile=roast_profile,
                origin=parse_origin(combined),
                price_aud=price,
                process=parse_process(combined, title=title, product_url=permalink),
                varietal=parse_varietal(combined),
                flavour_profile=parse_flavour_profile(combined),
                product_url=permalink,
                status="ok",
                error="",
            )
        )

    if not items:
        return [], "woocommerce_no_parsed_items"

    return items, None


def scrape_wordpress_product_api(
    session: requests.Session, roaster: str, listing_url: str
) -> Tuple[List[CoffeeItem], Optional[str]]:
    parsed = urlparse(listing_url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    endpoint = f"{base}/wp-json/wp/v2/product"

    all_products: List[Dict[str, object]] = []
    last_err: Optional[str] = None
    for page in range(1, 6):
        try:
            resp = session.get(
                endpoint,
                params={"per_page": 100, "page": page},
                timeout=TIMEOUT_SECONDS,
            )
            if resp.status_code != 200:
                last_err = f"wp_product_http_{resp.status_code}"
                break
            payload = resp.json()
            if not isinstance(payload, list):
                return [], "wp_product_invalid_payload"
            if not payload:
                break
            all_products.extend(payload)
            if len(payload) < 100:
                break
        except Exception as exc:
            return [], f"wp_product_error: {exc}"

    if not all_products:
        return [], last_err or "wp_product_no_items"

    items: List[CoffeeItem] = []
    for product in all_products:
        title_obj = product.get("title") or {}
        title = clean_text(str(title_obj.get("rendered", ""))) if isinstance(title_obj, dict) else ""
        product_url = clean_text(str(product.get("link", "")))
        excerpt_obj = product.get("excerpt") or {}
        content_obj = product.get("content") or {}
        excerpt = clean_text(BeautifulSoup(str(excerpt_obj.get("rendered", "")), "html.parser").get_text(" ")) if isinstance(excerpt_obj, dict) else ""
        content = clean_text(BeautifulSoup(str(content_obj.get("rendered", "")), "html.parser").get_text(" ")) if isinstance(content_obj, dict) else ""
        if not title or not product_url:
            continue
        if should_exclude_product_for_roaster(roaster, " ".join([title, product_url, excerpt])):
            continue
        combined = " ".join([title, excerpt, content])
        roast_profile = parse_roast_profile(combined)
        items.append(
            CoffeeItem(
                roaster=roaster,
                source_url=product_url,
                bean_name=title,
                roast_profile=roast_profile,
                origin=parse_origin(combined),
                price_aud=extract_price(combined),
                process=parse_process(combined, title=title, product_url=product_url),
                varietal=parse_varietal(combined),
                flavour_profile=parse_flavour_profile(combined),
                product_url=product_url,
                status="ok",
                error="",
            )
        )

    if not items:
        return [], "wp_product_no_parsed_items"
    return items, None


def scrape_one_roaster(session: requests.Session, roaster: Dict[str, object]) -> List[CoffeeItem]:
    name = roaster.get("name", "Unknown")
    url = str(roaster.get("url", "") or "").strip()
    collection_url_override = str(roaster.get("collection_url_override", "") or "").strip()
    shopify_collection_handle = str(roaster.get("shopify_collection_handle", "") or "").strip() or None
    price_variant_filter = str(roaster.get("price_variant_filter", "") or "").strip() or None
    force_html_listing = bool(roaster.get("force_html_listing", False))
    listing_url = collection_url_override or url

    if not listing_url:
        return [
            CoffeeItem(
                roaster=name,
                source_url="",
                bean_name="",
                roast_profile="",
                origin="",
                price_aud="",
                process="",
                varietal="",
                flavour_profile="",
                product_url="",
                status="error",
                error="missing_url",
            )
        ]

    listing_candidates = [listing_url]
    alt_url = alternate_www_url(listing_url)
    if alt_url and alt_url not in listing_candidates:
        listing_candidates.append(alt_url)

    candidate_errors: List[str] = []
    for candidate_url in listing_candidates:
        err = None
        discovery_errors: List[str] = []
        if not force_html_listing:
            items, err = scrape_shopify_collection_json(
                session,
                name,
                candidate_url,
                shopify_collection_handle=shopify_collection_handle,
                price_variant_filter=price_variant_filter,
            )
            if items:
                return items

            if not shopify_collection_handle:
                handles, discover_err = discover_collection_handles_from_listing(session, candidate_url)
                if discover_err:
                    discovery_errors.append(discover_err)
                best_items: List[CoffeeItem] = []
                best_handle: Optional[str] = None
                for handle in handles:
                    items, handle_err = scrape_shopify_collection_json(
                        session,
                        name,
                        candidate_url,
                        shopify_collection_handle=handle,
                    )
                    if items:
                        if len(items) > len(best_items):
                            best_items = items
                            best_handle = handle
                    if handle_err:
                        discovery_errors.append(f"{handle}:{handle_err}")
                if best_items:
                    if best_handle:
                        discovery_errors.append(f"selected_handle={best_handle}")
                    return best_items

        html_items, html_err = scrape_via_html_listing(session, name, candidate_url)
        if html_items:
            return html_items

        shopify_all_items, shopify_all_err = scrape_shopify_all_products_json(session, name, candidate_url)
        if shopify_all_items:
            return shopify_all_items

        shop_slug_items, shop_slug_err = scrape_via_shop_slug_listing(session, name, candidate_url)
        if shop_slug_items:
            return shop_slug_items

        sitemap_items, sitemap_err = scrape_via_sitemap(session, name, candidate_url)
        if sitemap_items:
            return sitemap_items

        woo_items, woo_err = scrape_woocommerce_store_api(session, name, candidate_url)
        if woo_items:
            return woo_items

        wp_items, wp_err = scrape_wordpress_product_api(session, name, candidate_url)
        if wp_items:
            return wp_items

        error_parts = [err, html_err, shopify_all_err, shop_slug_err, sitemap_err, woo_err, wp_err]
        if discovery_errors:
            error_parts.append("handle_discovery=" + "; ".join(discovery_errors))
        one_error_msg = " | ".join([e for e in error_parts if e]) or "unknown_scrape_error"
        candidate_errors.append(f"{candidate_url} => {one_error_msg}")

    error_msg = " || ".join(candidate_errors) if candidate_errors else "unknown_scrape_error"
    return [
        CoffeeItem(
            roaster=name,
            source_url=listing_url,
            bean_name="",
            roast_profile="",
            origin="",
            price_aud="",
            process="",
            varietal="",
            flavour_profile="",
            product_url="",
            status="error",
            error=error_msg,
        )
    ]


def item_identity(item: CoffeeItem) -> str:
    source = clean_text(item.source_url or item.product_url)
    if source:
        return source
    return f"{clean_text(item.roaster)}::{clean_text(item.bean_name)}"


def previous_item_identities(output_dir: Path) -> set:
    candidates = sorted(output_dir.glob("coffee_list_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        return set()

    try:
        payload = json.loads(candidates[0].read_text(encoding="utf-8"))
    except Exception:
        return set()

    items = payload.get("items", []) if isinstance(payload, dict) else []
    identities = set()
    for row in items:
        if not isinstance(row, dict):
            continue
        source = clean_text(str(row.get("source_url", "") or row.get("product_url", "")))
        if source:
            identities.add(source)
        else:
            identities.add(f"{clean_text(str(row.get('roaster', '')))}::{clean_text(str(row.get('bean_name', '')))}")
    return identities


def write_csv(path: Path, items: List[CoffeeItem]) -> None:
    def buy_link(url: str, label: str) -> str:
        u = clean_text(url)
        if not u:
            return ""
        text = clean_text(label) or "buy"
        text = text.replace('"', '""')
        return f'=HYPERLINK("{u}","{text}")'

    headers = [
        "roaster",
        "bean_name",
        "roast_profile",
        "origin",
        "price_aud",
        "process",
        "varietal",
        "flavour_profile",
        "status",
        "error",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for item in items:
            row = asdict(item)
            row["bean_name"] = buy_link(row.get("source_url", ""), row.get("bean_name", ""))
            row.pop("source_url", None)
            row.pop("product_url", None)
            writer.writerow(row)


def write_xlsx(path: Path, items: List[CoffeeItem], new_item_ids: set) -> None:
    headers = [
        "roaster",
        "bean_name",
        "roast_profile",
        "origin",
        "price_aud",
        "process",
        "varietal",
        "flavour_profile",
        "status",
        "error",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "coffee_list"
    ws.append(headers)

    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF59D")
    bean_name_max = len("bean_name")

    for item in items:
        row = asdict(item)
        bean_name = clean_text(str(row.get("bean_name", "")))
        if not bean_name and row.get("source_url"):
            bean_name = "buy"
        bean_name_max = max(bean_name_max, len(bean_name))
        values = [
            row.get("roaster", ""),
            bean_name,
            row.get("roast_profile", ""),
            row.get("origin", ""),
            row.get("price_aud", ""),
            row.get("process", ""),
            row.get("varietal", ""),
            row.get("flavour_profile", ""),
            row.get("status", ""),
            row.get("error", ""),
        ]
        ws.append(values)
        current_row = ws.max_row
        bean_cell = ws.cell(row=current_row, column=2)
        if row.get("source_url"):
            bean_cell.hyperlink = row["source_url"]
            bean_cell.style = "Hyperlink"

        if item.status == "ok" and item_identity(item) in new_item_ids:
            for col in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col).fill = yellow_fill

    ws.column_dimensions["B"].width = min(max(bean_name_max + 2, 14), 34)
    for col in ["A", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 18

    wb.save(path)


def write_json(path: Path, items: List[CoffeeItem], generated_at: str) -> None:
    records = [asdict(i) for i in items]
    grouped: Dict[str, List[Dict[str, str]]] = {
        "filter": [],
        "espresso": [],
        "omni": [],
        "unknown": [],
    }

    errors = []
    for row in records:
        if row["status"] == "error":
            errors.append(row)
        key = row["roast_profile"] if row["roast_profile"] in grouped else "unknown"
        grouped[key].append(row)

    payload = {
        "generated_at": generated_at,
        "summary": {
            "total_rows": len(records),
            "ok_rows": len([r for r in records if r["status"] == "ok"]),
            "error_rows": len(errors),
        },
        "grouped": grouped,
        "items": records,
        "errors": errors,
    }

    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def load_config(config_path: Path) -> List[Dict[str, object]]:
    with config_path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("Config must be a list of roaster objects")
    return data


def main() -> int:
    parser = argparse.ArgumentParser(description="Update local coffee bean list from roaster websites")
    parser.add_argument("--config", default=DEFAULT_CONFIG, help="Path to roaster config JSON")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Output directory")
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent
    config_path = (base_dir / args.config).resolve() if not Path(args.config).is_absolute() else Path(args.config)
    output_dir = (base_dir / args.output_dir).resolve() if not Path(args.output_dir).is_absolute() else Path(args.output_dir)

    if not config_path.exists():
        print(f"Config file not found: {config_path}", file=sys.stderr)
        return 1

    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        roasters = load_config(config_path)
    except Exception as exc:
        print(f"Failed to load config: {exc}", file=sys.stderr)
        return 1

    session = session_with_headers()
    all_items: List[CoffeeItem] = []
    previous_ids = previous_item_identities(output_dir)

    for roaster in roasters:
        all_items.extend(scrape_one_roaster(session, roaster))

    now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    csv_path = output_dir / f"coffee_list_{now}.csv"
    json_path = output_dir / f"coffee_list_{now}.json"
    xlsx_path = output_dir / f"coffee_list_{now}.xlsx"
    current_ok_ids = {item_identity(i) for i in all_items if i.status == "ok"}
    new_ids = current_ok_ids - previous_ids

    write_csv(csv_path, all_items)
    write_xlsx(xlsx_path, all_items, new_ids)
    write_json(json_path, all_items, generated_at=now)

    print(f"CSV saved: {csv_path}")
    print(f"XLSX saved: {xlsx_path}")
    print(f"JSON saved: {json_path}")
    print(f"Total rows: {len(all_items)}")
    print(f"Error rows: {len([i for i in all_items if i.status == 'error'])}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
